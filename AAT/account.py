import pip._vendor.requests as requests
import openpyxl
import psycopg2.errors
from pathlib import Path

from flask import (
    Blueprint, flash, g, redirect, render_template, request, url_for
)
from datetime import datetime, timezone
from werkzeug.exceptions import abort

from AAT.auth import login_required
from AAT.db import get_db

bp = Blueprint('account', __name__, url_prefix='/account')

# One is Monday

# def getPeriod(time):
#     teacher = 'teacherNone'
#     startTime = 540
#     if time>=500 and time<585:
#         teacher = 'teacher8'
#         startTime = 540
#     elif time>=1940 and time<2055: 
#         teacher = 'teacher1'
#         startTime = 1980
#     elif time>=4820 and time<4935:
#         teacher = 'teacher1'
#         startTime = 4860
#     elif time>=2055 and time<2145:
#         teacher = 'teacher2'
#         startTime = 2070
#     elif time>=4935 and time<5025:
#         teacher = 'teacher2'
#         startTime = 4950
#     elif time>=2175 and time<2265:
#         teacher = 'teacher3'
#         startTime = 2190
#     elif time>=5055 and time<5145:
#         teacher = 'teacher3'
#         startTime = 5070
#     elif time>=2265 and time<2355:
#         teacher = 'teacher7'
#         startTime = 2280
#     elif time>=5145 and time<5235:
#         teacher = 'teacher7'
#         startTime = 5160
#     elif time>=3460 and time<3555:
#         teacher = 'teacher4'
#         startTime = 3480
#     elif time>=6340 and time<6435:
#         teacher = 'teacher4'
#         startTime = 6360
#     elif time>=3555 and time<3645:
#         teacher = 'teacher5'
#         startTime = 3570
#     elif time>=6435 and time<6525:
#         teacher = 'teacher5'
#         startTime = 6450
#     elif time>=3675 and time<3765:
#         teacher = 'teacher6'
#         startTime = 3690
#     elif time>=6555 and time<6645:
#         teacher = 'teacher6'
#         startTime = 6570
#    return teacher, startTime

#transition week
# def getPeriod(time):
#     teacher = 'teacher8'
#     startTime = 1980
#     if time>1920 and time<2040:
#         startTime = 1980
#     elif time>3360 and time<3480:
#         startTime = 3420
#     elif time>4800 and time<4920:
#         startTime = 4860
#     elif time>6240 and time<6360:
#         startTime = 6300
#     return teacher, startTime

# hybrid schedule
def getPeriod(time):
    teacher = 'teacherNone'
    startTime = 540
    if time>=500 and time<585:
        teacher = 'teacher9'
        starTime = 540
    elif time>=1910 and time<2025:
        teacher = 'teacher1'
        startTime = 1950
    elif time>=2025 and time<2120:
        teacher = 'teacher2'
        startTime = 2045
    elif time>=2190 and time<2275:
        teacher = 'teacher3'
        startTime = 2200
    elif time>=2275 and time<2360:
        teacher = 'teacher7'
        startTime = 2285
    elif time>=3350 and time<3465:
        teacher = 'teacher1'
        startTime = 3390
    elif time>=3465 and time<3560:
        teacher = 'teacher2'
        startTime = 3485
    elif time>=3630 and time<3715:
        teacher = 'teacher3'
        startTime = 3640
    elif time>=3715 and time<3800:
        teacher = 'teacher7'
        startTime = 3725
    elif time>=4790 and time<4905:
        teacher = 'teacher4'
        startTime = 4830
    elif time>=4905 and time<5000:
        teacher = 'teacher5'
        startTime = 4925
    elif time>=5070 and time<5155:
        teacher = 'teacher6'
        startTime = 5080
    elif time>=5155 and time<5240:
        teacher = 'teacher0'
        startTime = 5165
    elif time>=6230 and time<6345:
        teacher = 'teacher4'
        startTime = 6270
    elif time>=6345 and time<6440:
        teacher = 'teacher5'
        startTime = 6365
    elif time>=6510 and time<6595:
        teacher = 'teacher6'
        startTime = 6520
    elif time>=6595 and time<6680:
        teacher = 'teacher0'
        startTime = 6605
    return teacher, startTime

#special mondays
# def getPeriod(time):
#     teacher = 'teacherNone'
#     startTime = 540
#     if time>=525 and time<585:
#         teacher = 'teacher8'
#         startTime = 540
#     elif time>=585 and time<675:
#         teacher = 'teacher4'
#         startTime = 600
#     elif time>=675 and time<765:
#         teacher = 'teacher5'
#         startTime = 690
#     elif time>=795 and time<885:
#         teacher = 'teacher6'
#         startTime = 810
#     return teacher, startTime

@bp.route('/scanner', methods=('GET',))
def scanner():
    return render_template('account/scanner.html')

@bp.route('/admit/<userID>', methods=('GET', 'POST'))
@login_required
def admit(userID):
    if request.method == 'POST':
        student = request.form['student']
        db_conn = get_db()
        db = db_conn.cursor()
        account = g.user
        db.execute('SELECT currentMeeting, joinTime FROM stranger WHERE userID = %s', (userID,))
        currentMeeting, joinTime = db.fetchone()
        if userID == None or userID == '':
            db.execute('UPDATE student SET currentMeeting = %s, joinTime = %s WHERE name = %s;', (currentMeeting, joinTime, student))
        else:
            db.execute('UPDATE student SET userID = %s, currentMeeting = %s, joinTime = %s, confidence = confidence+1 WHERE name = %s AND userID is NULL;', (userID, currentMeeting, joinTime, student))
            if db.rowcount <= 0:
                db.execute('UPDATE student SET currentMeeting = %s, joinTime = %s, confidence = confidence+1 WHERE name = %s AND userID = %s;', (currentMeeting, joinTime, student, userID))
                if db.rowcount <= 0:
                    #ID Mapping Collision
                    db.execute('UPDATE student SET currentMeeting = %s, joinTime = %s, confidence = confidence-1 WHERE name = %s;', (currentMeeting, joinTime, student))
            
        db.execute('DELETE FROM stranger WHERE userID = %s;', (userID,))
        db_conn.commit()
        return redirect(url_for('account.details'))

    elif request.method == 'GET':
        db_conn = get_db()
        db = db_conn.cursor()
        account = g.user
        db.execute('SELECT EXTRACT(DOW FROM joinTime), EXTRACT(HOUR FROM joinTime), EXTRACT(MINUTE FROM joinTime) FROM stranger WHERE userID = %s;', (userID,))
        time = db.fetchone()
        period = getPeriod((time[0]-1)*24*60+time[1]*60+time[2])[0]
        if period != 'teacherNone':
            db.execute('SELECT * FROM stranger WHERE userID = %s;', (userID,))
            stranger = db.fetchone()

            db.execute('SELECT name FROM student WHERE {} = %s AND currentMeeting != %s AND (userID is NULL OR confidence < 3);'.format(period), (account[0], stranger[1]))
            unassigned = db.fetchall()
            length = len(unassigned)
            #add new student
            if length == 0:
                #All students have been assigned. Overright
                db.execute('SELECT name FROM student WHERE {} = %s;'.format(period), (account[0],))
                return render_template('account/option.html', stranger=stranger[3], students=db.fetchall())
            elif length == 1:
                if stranger[0] == None or stranger[0] == '':
                    db.execute('UPDATE student SET currentMeeting = %s, joinTime = %s WHERE name = %s;', (stranger[1], stranger[2], unassigned[0][0]))
                else:
                    db.execute('UPDATE student SET userID = %s, currentMeeting = %s, confidence = 3, joinTime = %s WHERE name = %s;', (userID, stranger[1], stranger[2], unassigned[0][0]))
            else:
                return render_template('account/option.html', stranger=stranger[3], students=unassigned)
        
        db.execute('DELETE FROM stranger WHERE userID = %s;', (userID,))
        db_conn.commit()
        return redirect(url_for('account.details'))


@bp.route('/reject/<userID>', methods=('GET',))
@login_required
def reject(userID):
    db_conn = get_db()
    db = db_conn.cursor()
    db.execute('DELETE FROM stranger WHERE userID = %s;', (userID,))
    db_conn.commit()
    return redirect(url_for('account.details'))

@bp.route('/settings', methods=('GET', 'POST'))
@login_required
def settings():
    if request.method == 'POST':
        email = g.user[0]
        db_conn = get_db()
        db = db_conn.cursor()
        uploaded_file = request.files['upload-file']
        file_sheet = openpyxl.load_workbook(uploaded_file).active
        row_index = 1
        #reset roster
        if uploaded_file != None and file_sheet.max_row>10:
            db.execute('UPDATE student SET teacher0 = NULL WHERE teacher0 = %s;', (email,))
            db.execute('UPDATE student SET teacher1 = NULL WHERE teacher1 = %s;', (email,))
            db.execute('UPDATE student SET teacher2 = NULL WHERE teacher2 = %s;', (email,))
            db.execute('UPDATE student SET teacher3 = NULL WHERE teacher3 = %s;', (email,))
            db.execute('UPDATE student SET teacher4 = NULL WHERE teacher4 = %s;', (email,))
            db.execute('UPDATE student SET teacher5 = NULL WHERE teacher5 = %s;', (email,))
            db.execute('UPDATE student SET teacher6 = NULL WHERE teacher6 = %s;', (email,))
            db.execute('UPDATE student SET teacher7 = NULL WHERE teacher7 = %s;', (email,))
            db.execute('UPDATE student SET teacher8 = NULL WHERE teacher8 = %s;', (email,))
            db.execute('UPDATE student SET teacher9 = NULL WHERE teacher9 = %s;', (email,))

        while row_index < file_sheet.max_row:
            if file_sheet.cell(row=row_index, column=1).value=="Period":
                period = file_sheet.cell(row=row_index+1, column=1).value
                row_index+=4
                temp = row_index
                while file_sheet.cell(row=row_index, column=5).value!=None:
                    words = file_sheet.cell(row=row_index, column=5).value.replace('*','').split(',')
                    name = words[1].split(' ')[1]+" "+words[0]
                    cohort = file_sheet.cell(row=row_index, column=8).value
                    studentID = file_sheet.cell(row=row_index, column=2).value
                    db.execute('UPDATE student SET {} = %s, cohort = %s, studentID = %s WHERE name = %s AND {} is NULL;'.format("teacher"+period, "teacher"+period), (email, cohort, studentID, name))
                    if db.rowcount <= 0:
                        #check for unique error
                        try:
                            db.execute(
                                'INSERT INTO student (currentMeeting, name, {}, cohort, studentID, confidence) VALUES (%s, %s, %s, %s, %s, %s);'.format("teacher"+period),
                                ("area51", name, email, cohort, studentID, 0)
                            )
                        except psycopg2.errors.UniqueViolation:
                            flash('Student {} already has a teacher for Period {}'.format(name,period))
                            return render_template('account/settings.html')
                    row_index += 1
            else:
                row_index += 1
        db_conn.commit()
        return redirect(url_for('account.details'))
    
    return render_template('account/settings.html')

@bp.route('/removeAbsence/<string:name>/<boolean>', methods=('GET',))
@login_required
def removeAbsence(name, boolean):
    db_conn = get_db()
    db = db_conn.cursor()
    account = g.user
    db.execute('SELECT cohort, currentMeeting FROM student WHERE name = %s;', (name,))
    student = db.fetchone()
    cohort = student[0]
    if boolean == 'perm':
        if 'C' in cohort:
            #completely remote but change to in person
            db.execute('UPDATE student SET cohort = %s WHERE name = %s;', (cohort[:-1], name))
        else:
            #in person but change to on zoom
            db.execute('UPDATE student SET cohort = %s, currentMeeting = %s, joinTime = %s WHERE name = %s;', (cohort+'C', account[2], account[3], name))
    elif boolean == 'temp':
        if 'C' in cohort:
            #completely remote change to in person
            db.execute('UPDATE student SET currentMeeting = %s WHERE name = %s;', ('temporary', name))
        else:
            #in person but change to on zoom
            db.execute('UPDATE student SET currentMeeting = %s WHERE name = %s;', ('temporary', name))
    db_conn.commit()
    return redirect(url_for('account.details'))

# bp.route('/removeAbsence/<name>', methods=('POST',))
# @login_required
# def removeAbsence(name):
#     db_conn = get_db()
#     db = db_conn.cursor()
#     account = g.user
#     db.execute('SELECT cohort, currentMeeting FROM student WHERE name = %s;', (name,))
#     student = db.fecthone()
#     cohort = student[0]
#     if 'C' in cohort:
#         #completely remote change to in person
#         db.execute('UPDATE student SET currentMeeting = %s WHERE name = %s;', (account[2], name))
#     else:
#         #in person but change to on zoom
#         db.execute('UPDATE student SET cohort = %s, currentMeeting = %s, joinTime = %s WHERE name = %s;', (cohort+'C', account[2], account[3], name))
#     return redirect(url_for('account.details'))

@bp.route('/meetingReset', methods=('GET',))
@login_required
def meetingReset():
    db_conn = get_db()
    db = db_conn.cursor()
    account = g.user
    userID = account[4]

    db.execute('SELECT name FROM stranger WHERE currentMeeting = %s;', (account[2],))
    stranger = [unknown[0] for unknown in db]
    
    db.execute('DELETE FROM stranger WHERE currentMeeting = %s;', (account[2],))

    db.execute('SELECT EXTRACT(DOW FROM %s), EXTRACT(HOUR FROM %s), EXTRACT(MINUTE FROM %s);', (account[3], account[3], account[3]))
    time = db.fetchone()
    period, startTime = getPeriod((time[0]-1)*24*60+time[1]*60+time[2])
    if period == 'teacherNone':
        db.execute('UPDATE teacher SET currentMeeting = NULL WHERE userID = %s;', (userID,))
        db_conn.commit()
        return redirect(url_for('account.details'))
    # assert that meeting is a class
    
    current_time = datetime.now(timezone.utc)
    meeting_duration = current_time - account[3].astimezone(timezone.utc)
    if meeting_duration.total_seconds() < 180:
        db.execute('UPDATE teacher SET currentMeeting = NULL WHERE userID = %s;', (userID,))
        db_conn.commit()
        return redirect(url_for('account.details'))
    # assert that meeting duration is more than 3 minutes (not a misclick)

    db.execute('SELECT COUNT(*) FROM history WHERE teacher = %s AND period = %s AND EXTRACT(DAY FROM date) = %s;', (account[0], period, current_time.day))
    if db.fetchone()[0] > 0:
        db.execute('UPDATE teacher SET currentMeeting = NULL WHERE userID = %s;', (userID,))
        db_conn.commit()
        return redirect(url_for('account.details'))
    # #assert that only one meeting per period per day

    db.execute('SELECT EXTRACT(DOW FROM joinTime), EXTRACT(HOUR FROM joinTime), EXTRACT(MINUTE FROM joinTime), name, currentMeeting, cohort FROM student WHERE {} = %s;'.format(period), (account[0],))
    roster = db.fetchall()
    if len(roster) == 0:
        db.execute('UPDATE teacher SET currentMeeting = NULL WHERE userID = %s;', (userID,))
        db_conn.commit()
        return redirect(url_for('account.details'))
    #assert that there are students in the class
    
    tardy = []
    absent = []

    for student in roster:
        name = student[3]
        currentMeeting = student[4]
        cohort = student[5]
        if cohort is not None and 'C' in cohort:
            #student is completely remote
            if currentMeeting == account[2] and student[0] != None:
                #student is in zoom meeting
                joinTime = (student[0]-1)*24*60+student[1]*60+student[2]
                if joinTime-startTime>30:
                    absent.append(name)
                elif joinTime-startTime>account[5]:
                    tardy.append(name)
            elif currentMeeting != 'temporary':
                absent.append(name)
        elif cohort == 'A':
            #cohort A
            if time[0] == 3 or time[0] == 5:
                #is expected to be there
                if currentMeeting == account[2] and student[0] != None:
                    #student is in zoom meeting
                    joinTime = (student[0]-1)*24*60+student[1]*60+student[2]
                    if joinTime-startTime>30:
                        absent.append(name)
                    elif joinTime-startTime>account[5]:
                        tardy.append(name)
                else:
                    absent.append(name)
            else:
                #not supposed to be there so stranger
                if currentMeeting == account[2] and currentMeeting != 'temporary':
                    stranger.append(name)
        elif cohort == 'B':
            #cohort B
            if time[0] == 2 or time[0] == 4:
                #is expected to be there
                if currentMeeting == account[2] and student[0] != None:
                    #student is in zoom meeting
                    joinTime = (student[0]-1)*24*60+student[1]*60+student[2]
                    if joinTime-startTime>30:
                        absent.append(name)
                    elif joinTime-startTime>account[5]:
                        tardy.append(name)
                else:
                    absent.append(name)
            else:
                #not supposed to be there so stranger
                if currentMeeting == account[2] and currentMeeting != 'temporary':
                    stranger.append(name)
    
    #add to history
    db.execute('UPDATE history SET absent = %s, tardy = %s, stranger = %s, date = %s WHERE period = %s AND teacher = %s;', (absent, tardy, stranger, account[3], period, account[0]))
    if db.rowcount <= 0:
        db.execute('INSERT INTO history (absent, tardy, stranger, date, period, teacher) VALUES (%s, %s, %s, %s, %s, %s);', (absent, tardy, stranger, account[3], period, account[0]))

    db.execute('UPDATE teacher SET startTime = NOW(), currentMeeting = %s WHERE email = %s;', ('reset', account[0]))
    db_conn.commit()

    return redirect(url_for('account.details'))

@bp.route('/history', methods=('GET',))
@login_required
def history():
    db = get_db().cursor()
    account = g.user

    db.execute('SELECT absent, tardy, stranger, period FROM history WHERE teacher = %s;', (account[0],))
    return render_template('account/history.html', absentList= db.fetchall())

@bp.route('/roster', methods=('GET', 'POST'))
@login_required
def roster():
    db = get_db().cursor()
    account = g.user
    db.execute('SELECT EXTRACT(DOW FROM startTime), EXTRACT(HOUR FROM startTime), EXTRACT(MINUTE FROM startTime) FROM teacher WHERE email = %s;', (account[0],))
    time = db.fetchone()
    period, startTime = getPeriod((time[0]-1)*24*60+time[1]*60+time[2])
    if period == 'teacherNone':
        period = 'teacher9'
    
    db.execute('SELECT name FROM student WHERE {} = %s AND currentMeeting = %s;'.format(period), (account[0], account[2]))
    roster = db.fetchall()
    return render_template('account/roster.html', roster=roster, period=period[-1:])

@bp.route('/dashboard', methods=('GET', 'POST'))
@login_required
def details():
    if request.method == 'POST':
        db_conn = get_db()
        db = db_conn.cursor()
        account = g.user
        tardyTime = request.form['tardy']
        if len(tardyTime) > 0:
            db.execute('UPDATE teacher SET tardyTime = %s WHERE email = %s;', (tardyTime, account[0]))
            db_conn.commit()
        return redirect(url_for('account.details'))

    elif request.method == 'GET':
        db = get_db().cursor()
        account = g.user
        teacherMeeting = account[2]

        db.execute('SELECT EXTRACT(DOW FROM startTime), EXTRACT(HOUR FROM startTime), EXTRACT(MINUTE FROM startTime) FROM teacher WHERE email = %s;', (account[0],))
        time = db.fetchone()
        currentTime = (time[0]-1)*24*60+time[1]*60+time[2]
        period, startTime = getPeriod(currentTime)


        student_list = []
        if period != 'teacherNone':
            db.execute('SELECT EXTRACT(DOW FROM joinTime), EXTRACT(HOUR FROM joinTime), EXTRACT(MINUTE FROM joinTime), name, currentMeeting, cohort FROM student WHERE {} = %s;'.format(period), (account[0],))
            student_list = db.fetchall()
        else:
            period = 'teacher9'
        
        if teacherMeeting == 'reset' and len(student_list) > 0:
            meeting_counts = {}
            for student in student_list:
                if student[0] == time[0]:
                    meeting = student[4]
                    if meeting in meeting_counts:
                        meeting_counts[meeting] += 1
                    else:
                        meeting_counts[meeting] = 1
            
            max = 0
            maybeMeeting = None
            for key in meeting_counts.keys():
                if meeting_counts[key] > max:
                    max = meeting_counts[key]
                    maybeMeeting = key
            
            teacherMeeting = maybeMeeting
            if max >= 3:
                db.execute('UPDATE teacher SET currentMeeting = %s WHERE email = %s;', (maybeMeeting, account[0]))
                get_db().commit()
            

        if teacherMeeting == None or len(student_list) <= 0:
            #look in history if no class is currently live
            classes = ['teacher6', 'teacher5', 'teacher4', 'teacher7', 'teacher3', 'teacher2', 'teacher1', 'teacher6', 'teacher5', 'teacher4']
            temp = False
            for c in classes:
                if c == period:
                    temp = True

                if temp:
                    db.execute('SELECT absent, tardy, stranger FROM history WHERE period = %s AND teacher = %s;', (c, account[0]))
                    if db.rowcount > 0:
                        #history entry found
                        absent, tardy, stranger = db.fetchone()
                        return render_template('account/dashboard.html', startTime=startTime, period=period[-1:], absent=absent, tardy=tardy, stranger=stranger, name=account[6], numPresent=0, tardyTime=account[5])
            
            #no history entries exist
            return render_template('account/dashboard.html', startTime=startTime, period=period[-1:], absent=[], tardy=[], stranger=[], name=account[6], numPresent=0, tardyTime=account[5])
        
        tardy = []
        absent = []
        stranger_danger = []
        numPresent = 0

        for student in student_list:
            name = student[3]
            currentMeeting = student[4]
            cohort = student[5]
            if cohort is not None and 'C' in cohort:
                #student is completely remote
                if currentMeeting == teacherMeeting and student[0] != None:
                    #student is in zoom meeting
                    numPresent += 1
                    joinTime = (student[0]-1)*24*60+student[1]*60+student[2]
                    if joinTime-startTime>30:
                        absent.append((name,cohort))
                    elif joinTime-startTime>account[5]:
                        tardy.append(name)
                elif currentMeeting != 'temporary':
                    absent.append((name,cohort))
            elif cohort == 'A':
                #cohort A
                if time[0] == 3 or time[0] == 5:
                    #is expected to be there
                    if currentMeeting == teacherMeeting and student[0] != None:
                        #student is in zoom meeting
                        numPresent += 1
                        joinTime = (student[0]-1)*24*60+student[1]*60+student[2]
                        if joinTime-startTime>30:
                            absent.append((name,cohort))
                        elif joinTime-startTime>account[5]:
                            tardy.append(name)
                    else:
                        absent.append((name,cohort))
                else:
                    #not supposed to be there so stranger
                    if currentMeeting == teacherMeeting:
                        numPresent += 1
                        if currentMeeting != 'temporary':
                            stranger_danger.append((name, None))
            elif cohort == 'B':
                #cohort B
                if time[0] == 2 or time[0] == 4:
                    #is expected to be there
                    if currentMeeting == teacherMeeting and student[0] != None:
                        #student is in zoom meeting
                        numPresent += 1
                        joinTime = (student[0]-1)*24*60+student[1]*60+student[2]
                        if joinTime-startTime>30:
                            absent.append((name,cohort))
                        elif joinTime-startTime>account[5]:
                            tardy.append(name)
                    else:
                        absent.append((name,cohort))
                else:
                    #not supposed to be there so stranger
                    if currentMeeting == teacherMeeting:
                        numPresent += 1
                        if currentMeeting != 'temporary':
                            stranger_danger.append((name, None))

                
        db.execute('SELECT name, userID FROM stranger WHERE currentMeeting = %s;', (teacherMeeting,))
        strangers = list(db.fetchall())
        stranger_danger += strangers
        numPresent += len(strangers)

        return render_template('account/dashboard.html', startTime=0, period=period[-1:], absent=absent, tardy=tardy, stranger=stranger_danger, name=account[6], numPresent=numPresent, tardyTime=account[5])